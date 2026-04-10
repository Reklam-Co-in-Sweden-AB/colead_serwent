"""
Importscript för Komtek tømningshistorik till Serwent-plattformen.

Läser en Excel-fil exporterad från Komtek och importerar tömningsdata
grupperat per sone (Rodenavn) och vecka till serwent_produksjon.

Kolumn U (Rodenavn) mappas till soner i systemet.

Kör: python3 scripts/import-komtek.py <sökväg-till-excel> [kommun]
"""
import sys
import json
import urllib.request
import openpyxl

# Konfiguration
API_URL = "https://colead-serwent.vercel.app/api/komtek/import"
DEFAULT_KOMMUNE = "Vestre Toten"

# Kolumnordning i Komtek-export
COLUMNS = [
    "kommune",        # A
    "tomme_dato",     # B
    "anlegg",         # C
    "kunde",          # D
    "adresse",        # E
    "postnummer",     # F
    "poststed",       # G
    "eiendom",        # H
    "anleggstype",    # I
    "avtaletype",     # J
    "type_tomming",   # K
    "bil",            # L
    "adkomst",        # M
    "tomt",           # N
    "tomme_volum",    # O
    "slangeutlegg",   # P
    "hoydeforskjell", # Q
    "tommer",         # R
    "merknad",        # S
    "avvik",          # T
    "rodenavn",       # U
]


def parse_excel(path: str) -> list[dict]:
    """Läs Komtek Excel-fil och returnera lista med rader."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    rows_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not any(row):
            continue  # Hoppa över tomma rader

        entry = {}
        for i, col_name in enumerate(COLUMNS):
            val = row[i] if i < len(row) else None

            # Konvertera datum till ISO-format
            if col_name == "tomme_dato" and val is not None:
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                else:
                    val = str(val)

            # Konvertera numeriska fält
            if col_name in ("tomme_volum", "slangeutlegg", "hoydeforskjell"):
                if val is not None:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = None

            # Övriga fält som sträng
            if val is not None and col_name not in ("tomme_volum", "slangeutlegg", "hoydeforskjell", "tomme_dato"):
                val = str(val).strip()

            entry[col_name] = val

        # Hoppa över rader utan Rodenavn eller datum
        if not entry.get("rodenavn") or not entry.get("tomme_dato"):
            continue

        # Hoppa över header-raden som kan ha hamnat bland data
        if entry["rodenavn"] == "Rodenavn":
            continue

        rows_data.append(entry)

    return rows_data


def main():
    if len(sys.argv) < 2:
        print("Användning: python3 scripts/import-komtek.py <excel-fil> [kommun]")
        print("Exempel:    python3 scripts/import-komtek.py ~/Downloads/TømmeRapport.xlsx 'Vestre Toten'")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    kommune = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_KOMMUNE

    print(f"Läser fil: {xlsx_path}")
    rows = parse_excel(xlsx_path)
    print(f"  → {len(rows)} giltiga rader parsade")

    if not rows:
        print("Inga rader att importera. Kontrollera filen.")
        sys.exit(1)

    # Visa sammanfattning per Rodenavn
    from collections import Counter
    sone_counts = Counter(r["rodenavn"] for r in rows)
    print(f"\nKommun: {kommune}")
    print(f"Soner hittade ({len(sone_counts)}):")
    for sone, count in sone_counts.most_common():
        print(f"  {sone}: {count} tömningar")

    # Visa datumintervall
    dates = sorted(r["tomme_dato"] for r in rows if r["tomme_dato"])
    print(f"\nPeriod: {dates[0][:10]} — {dates[-1][:10]}")

    # Bekräfta
    secret = input("\nKlistra in SUPABASE_SERVICE_ROLE_KEY: ").strip()
    if not secret:
        print("Avbryter — nyckel saknas.")
        sys.exit(1)

    confirm = input(f"\nImportera {len(rows)} tömningar till {kommune}? (ja/nej): ").strip()
    if confirm.lower() != "ja":
        print("Avbryter.")
        sys.exit(0)

    # Skicka till API
    payload = json.dumps({"rows": rows, "kommune": kommune}).encode("utf-8")

    req = urllib.request.Request(
        API_URL,
        data=payload,
        headers={
            "Content-Type": "application/json",
            "x-import-secret": secret,
        },
    )

    try:
        with urllib.request.urlopen(req) as res:
            result = json.loads(res.read())
    except Exception as e:
        print(f"\nFEL vid API-anrop: {e}")
        sys.exit(1)

    # Visa resultat
    print(f"\n{'='*50}")
    print(f"Import klar!")
    print(f"  Totalt rader:        {result.get('total_rows', 0)}")
    print(f"  Hoppade över:        {result.get('skipped', 0)}")
    print(f"  Aggregerade veckor:  {result.get('aggregated_weeks', 0)}")
    print(f"  Upsertade:           {result.get('upserted', 0)}")

    if result.get("created_soner"):
        print(f"\n  Nya soner skapade ({len(result['created_soner'])}):")
        for s in result["created_soner"]:
            print(f"    + {s}")

    if result.get("errors"):
        print(f"\n  Fel ({len(result['errors'])}):")
        for err in result["errors"][:10]:
            print(f"    - {err}")


if __name__ == "__main__":
    main()
